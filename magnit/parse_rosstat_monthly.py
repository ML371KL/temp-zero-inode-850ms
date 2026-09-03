"""Parse Rosstat monthly food CPI (sheet 02) 2024-2026 -> quarterly cumulative food inflation.
Official anchor to validate the weekly proxy against.
"""
import openpyxl, json, pathlib

SRC = pathlib.Path(__file__).parent / "data" / "macro" / "rosstat_monthly_ipc.xlsx"
OUT = pathlib.Path(__file__).parent / "data" / "macro"
MON = ["январь", "февраль", "март", "апрель", "май", "июнь",
       "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["02"]
years = [ws.cell(4, c).value for c in range(2, ws.max_column + 1)]
print("year cols:", years[-6:])
col = {y: 2 + i for i, y in enumerate(years)}
monthly = {}
for r in range(6, 18):
    m = ws.cell(r, 1).value
    for y in [y for y in years if isinstance(y, int) and y >= 2019]:
        v = ws.cell(r, col[y]).value
        if v is None: continue
        monthly[f"{y}-{MON.index(m)+1:02d}"] = float(v) - 100
q = {}
for k, v in sorted(monthly.items()):
    y, m = k.split("-")
    qq = f"{y}-Q{(int(m)-1)//3+1}"
    q.setdefault(qq, []).append(v)
qcumul = {}
for k, vs in sorted(q.items()):
    p = 1.0
    for x in vs: p *= (1 + x / 100)
    qcumul[k] = round((p - 1) * 100, 2)
(OUT / "food_monthly.json").write_text(json.dumps(
    {"monthly_food_mom_pct": {k: round(v, 2) for k, v in sorted(monthly.items())},
     "quarterly_food_cumul_pct": qcumul,
     "source": "rosstat ipc_mes_07-2026.xlsx sheet 02 (food m/m)", "method": "chain-linked m/m"},
    ensure_ascii=False, indent=1), encoding="utf-8")
for k in ("2024-Q1", "2024-Q4", "2025-Q1", "2025-Q4", "2026-Q1", "2026-Q2", "2026-Q3"):
    print(f"  {k}: {qcumul.get(k)}%  n_months={len(q.get(k, []))}")
