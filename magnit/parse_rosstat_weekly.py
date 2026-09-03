"""Parse Rosstat weekly CPI (nedel_Ipc.xlsx) -> food-basket (rows 0-43: beef..vodka) weekly inflation.
Chain-links weekly % into quarterly averages. Saves magnit/data/macro/food_weekly.json.
Method: equal-weighted mean of 44 food items (proxy; official CPI uses expenditure weights -> flagged).
Freshness: weekly, Wednesday release, lag ~3 days (file dated 02.09.2026 with week ending 31.08).
"""
import openpyxl, json, pathlib, datetime, re, statistics

SRC = pathlib.Path(__file__).parent / "data" / "macro" / "rosstat_weekly_ipc.xlsx"
OUT = pathlib.Path(__file__).parent / "data" / "macro"
N_FOOD = 44  # rows 0..43
MONTHS_RU = {"января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5, "июня": 6,
             "июля": 7, "августа": 8, "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12}

def parse_date(hdr, year):
    m = re.search(r"на (\d{1,2}) (\S+)", str(hdr))
    if not m: return None
    mon = MONTHS_RU.get(m.group(2).lower(), 1)
    d = datetime.date(year, mon, int(m.group(1)))
    if d.month == 1 and int(m.group(1)) > 20 and year > 2022:
        pass  # early-Jan belongs to this year sheet already
    return d

def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    weeks = []  # (date, food_weekly_pct)
    for sh in ("2024", "2025", "2026"):
        ws = wb[sh]
        year = int(sh)
        for c in range(2, ws.max_column + 1):
            d = parse_date(ws.cell(4, c).value, year)
            if d is None: continue
            vals = []
            for r in range(5, 5 + N_FOOD):
                v = ws.cell(r, c).value
                try: vals.append(float(v))
                except (TypeError, ValueError): pass
            if len(vals) >= 40:
                vals.sort()
                k = len(vals) // 10  # 10-90 trimmed mean: drops seasonal veg spikes
                trim = vals[k:len(vals) - k] if k else vals
                med = statistics.median(vals)
                weeks.append((d.isoformat(), round(sum(trim) / len(trim) - 100, 3),
                              round(med - 100, 3), len(vals)))
    weeks.sort()
    # quarterly mean weekly % and approx cumulative (chain) for BOTH estimators
    from collections import defaultdict
    q = defaultdict(list)
    qmed = defaultdict(list)
    for d, pct, med, n in weeks:
        dt = datetime.date.fromisoformat(d)
        key = f"{dt.year}-Q{(dt.month-1)//3+1}"
        q[key].append(pct); qmed[key].append(med)
    def cumul(vs):
        p = 1.0
        for x in vs: p *= (1 + x / 100)
        return round((p - 1) * 100, 3)
    qavg = {k: round(sum(v) / len(v), 4) for k, v in sorted(q.items())}
    qcumul = {k: cumul(v) for k, v in sorted(q.items())}
    qmed_cumul = {k: cumul(v) for k, v in sorted(qmed.items())}
    # validate vs official monthly anchor
    official = {}
    try:
        official = json.loads((OUT / "food_monthly.json").read_text(encoding="utf-8"))["quarterly_food_cumul_pct"]
    except FileNotFoundError:
        pass
    (OUT / "food_weekly.json").write_text(json.dumps(
        {"weeks": [{"date": d, "food_wow_trimmed_pct": p, "food_wow_median_pct": m, "n_items": n} for d, p, m, n in weeks],
         "quarterly_trimmed_cumul_pct": qcumul, "quarterly_median_cumul_pct": qmed_cumul,
         "official_quarterly_food_cumul_pct": official,
         "method": "44 food items; trimmed 10-90 mean + median; official monthly CPI for anchors",
         "source": "https://rosstat.gov.ru/storage/mediabank/nedel_Ipc.xlsx",
         "n_weeks": len(weeks), "range": [weeks[0][0], weeks[-1][0]]},
        ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"weeks: {len(weeks)} range {weeks[0][0]}..{weeks[-1][0]}")
    print(f"{'q':<9} {'trimmed':>9} {'median':>9} {'official':>9}")
    for k in ("2024-Q1", "2024-Q4", "2025-Q1", "2025-Q4", "2026-Q1", "2026-Q2"):
        print(f"  {k}: {qcumul.get(k, float('nan')):>+8.2f}% {qmed_cumul.get(k, float('nan')):>+8.2f}% {official.get(k, float('nan')):>+8}%")

if __name__ == "__main__":
    main()
