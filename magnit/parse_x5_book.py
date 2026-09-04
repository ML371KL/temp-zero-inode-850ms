"""Parse X5 analyst databook (IAS 17 = pre16-equivalent basis, same as Magnit pre16).
Sheets: P&L (revenue/EBITDA levels), EBITDA (margins), Debt (net debt pre16 + ND/EBITDA),
Operating Results (yoy growth + LFL split). Quarterly Q1 2023-Q2 2026 + annual 2021-2025.
Saves magnit/data/peers/x5_quarterly.json
"""
import openpyxl, json, pathlib, re

SRC = pathlib.Path(__file__).parent / "data" / "peers" / "financial_and_operating_results_q2_2026.xlsx"
OLD = pathlib.Path(__file__).parent / "data" / "peers" / "financial_and_operating_results_q1_2024.xlsx"

def rev_old_2022():
    wb0 = openpyxl.load_workbook(OLD, data_only=True)
    ws0 = wb0["Profit and Loss"]
    c2022 = {}
    for c in range(1, ws0.max_column + 1):
        h = ws0.cell(5, c).value
        if isinstance(h, str) and re.fullmatch(r"Q[1-4] 20(1[89]|2[0-3])", h.strip()):
            c2022[h.strip()] = c
    r0 = next(r for r in range(1, ws0.max_row + 1) if str(ws0.cell(r, 2).value).strip() == "Revenue")
    return {f"20{h[-2:]}Q{h[1]}": ws0.cell(r0, c).value for h, c in c2022.items()}
wb = openpyxl.load_workbook(SRC, data_only=True)

def cols_ias17_quarterly(ws):
    out = {}
    for c in range(1, ws.max_column + 1):
        if ws.cell(4, c).value == "IAS 17" and isinstance(ws.cell(5, c).value, str) and "КВ." in ws.cell(5, c).value:
            out[ws.cell(5, c).value.strip()] = c
    return out

QN = {"1 КВ. 2023": "2023Q1", "2 КВ. 2023": "2023Q2", "3 КВ. 2023": "2023Q3", "4 КВ. 2023": "2023Q4",
      "1 КВ. 2024": "2024Q1", "2 КВ. 2024": "2024Q2", "3 КВ. 2024": "2024Q3", "4 КВ. 2024": "2024Q4",
      "1 КВ. 2025": "2025Q1", "2 КВ. 2025": "2025Q2", "3 КВ. 2025": "2025Q3", "4 КВ. 2025": "2025Q4",
      "1 КВ. 2026": "2026Q1", "2 КВ. 2026": "2026Q2"}

def rows_by_label(ws):
    return {str(ws.cell(r, 2).value).strip(): r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value}

pl, eb, db, op = wb["Profit and Loss"], wb["EBITDA"], wb["Debt"], wb["Operating Results"]
qc = cols_ias17_quarterly(pl)
rl, re_, rd = rows_by_label(pl), rows_by_label(eb), rows_by_label(db)
out = {"quarters": {}, "annual": {}}
prev_rev = {}
for hdr, c in sorted(qc.items(), key=lambda x: x[0]):
    q = QN[hdr]
    rev = pl.cell(rl["Выручка"], c).value
    ebitda = eb.cell(re_["EBITDA"], c).value if "EBITDA" in re_ else None
    ebitda_m = None
    for k in re_:
        if k.startswith("Рентабельность EBITDA"):
            ebitda_m = eb.cell(re_[k], c).value; break
    out["quarters"][q] = {"revenue_mln": rev, "ebitda_mln": ebitda,
                          "ebitda_margin": round(ebitda_m * 100, 2) if isinstance(ebitda_m, float) and ebitda_m < 1 else ebitda_m}
# yoy from levels (2023 via old book: identical Q1 2023 in both perimeters -> no break)
qs = sorted(out["quarters"])
for i, q in enumerate(qs):
    y, qq = int(q[:4]), q[4:]
    pq = f"{y-1}{qq}"
    base = out["quarters"].get(pq, {}).get("revenue_mln")
    if base is None and pq.startswith("2022"):
        base = rev_old_2022().get(pq)
    cur = out["quarters"][q]["revenue_mln"]
    if base and cur:
        out["quarters"][q]["revenue_yoy"] = round((cur / base - 1) * 100, 1)
# annual IAS17
for c in range(1, pl.max_column + 1):
    if pl.cell(4, c).value == "IAS 17" and isinstance(pl.cell(5, c).value, int):
        y = str(pl.cell(5, c).value)
        if y in ("2021", "2022", "2023", "2024", "2025"):
            rev = pl.cell(rl["Выручка"], c).value
            out["annual"][y] = {"revenue_mln": rev}
# debt: quarterly balances
dq = {}
for c in range(1, db.max_column + 1):
    v = db.cell(5, c).value
    if v and "00:00" in str(v):
        dq[str(v)[:10]] = c
for d, c in sorted(dq.items()):
    nd = db.cell(rd["Чистый долг до применения МСФО (IFRS) 16"], c).value
    ndx = db.cell(rd["Чистый долг / EBITDA до применения МСФО (IFRS) 16"], c).value
    out.setdefault("debt_quarters", {})[d] = {"net_debt_pre16_mln": nd, "nd_ebitda_pre16": ndx}
# LFL from Operating Results: rows Продажи/Трафик/Средний чек under 'LFL - Итого' (fractions)
op_labels = rows_by_label(op)
print("op labels with LFL:", [k for k in op_labels if "LFL" in k][:10])
# quarterly headers in op sheet: row5 cols with 'КВ.'
oq = {}
for c in range(1, op.max_column + 1):
    v = op.cell(5, c).value
    if isinstance(v, str) and "КВ." in v:
        oq[v.strip()] = c
print("op quarter cols:", oq)
QN_OP = {"2 КВ. 2024": "2024Q2", "3 КВ. 2024": "2024Q3", "4 КВ. 2024": "2024Q4",
         "1 КВ. 2025": "2025Q1", "2 КВ. 2025": "2025Q2", "3 КВ. 2025": "2025Q3",
         "4 КВ. 2025": "2025Q4", "1 КВ. 2026": "2026Q1", "2 КВ. 2026": "2026Q2"}
r_lfl = op_labels["LFL - Итого"]
lab = {str(op.cell(r, 2).value).strip(): r for r in range(r_lfl, r_lfl + 5)}
for hdr, c in oq.items():
    if hdr not in QN_OP: continue
    q = QN_OP[hdr]
    out["quarters"].setdefault(q, {}).update({
        "x5_lfl": round(op.cell(lab["Продажи"], c).value * 100, 1),
        "x5_lfl_traffic": round(op.cell(lab["Трафик"], c).value * 100, 1),
        "x5_lfl_ticket": round(op.cell(lab["Средний чек"], c).value * 100, 1)})
out["x5_old_perimeter_mln"] = rev_old_2022()
out["x5_old_perimeter_2022_mln"] = {k: v for k, v in out["x5_old_perimeter_mln"].items() if k.startswith("2022")}
pathlib.Path("magnit/data/peers/x5_quarterly.json").write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
for q in qs: print(f"  {q}: rev={out['quarters'][q]['revenue_mln']} yoy={out['quarters'][q].get('revenue_yoy')} ebitda_m={out['quarters'][q]['ebitda_margin']}")
print("annual:", {k: v["revenue_mln"] for k, v in out["annual"].items()})
